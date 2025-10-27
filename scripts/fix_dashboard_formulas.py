from openpyxl import load_workbook

BOARD_RANGE_ROWS = 600
START_ROW = 6

COLUMN_FORMULAS = {
    "I": '=IF($H{row}="","",IFERROR(RssMarket($H{row},"銘柄名称"),""))',
    "J": '=IF(OR($N{row}="",$O{row}="",$AA{row}=0),"",(($N{row}-$O{row})/$AA{row})/100)',
    "K": '=IF(OR($AD{row}="",$J{row}="",$AD{row}=0),"",ABS($J{row}-$AD{row})/ABS($AD{row})*100)',
    "L": '=""',
    "M": '=""',
    "N": '=IF($H{row}="","",IFERROR(RssMarket($H{row},"現在値"),""))',
    "O": '=IF($H{row}="","",IFERROR(RssMarket($H{row},"出来高加重平均"),""))',
    "P": '=IF($H{row}="","",IFERROR(RssMarket($H{row},"前日終値"),""))',
    "Q": '=IF($H{row}="","",IFERROR(RssMarket($H{row},"56"),""))',
    "R": '=IF($H{row}="","",IFERROR(RssMarket($H{row},"55"),""))',
    "S": '=IF(OR(Q{row}="",R{row}=""),"",(Q{row}+R{row})/2)',
    "T": '=IF(OR(S{row}="",P{row}=""),"",(S{row}-P{row})/P{row}*10000)',
}


def apply_formulas(ws):
    for col, template in COLUMN_FORMULAS.items():
        for offset in range(BOARD_RANGE_ROWS):
            row = START_ROW + offset
            ws[f"{col}{row}"] = template.format(row=row)


def main() -> None:
    path = r"C:\AI\asagake\SHINSOKU.xlsm"
    wb = load_workbook(path)
    ws = wb["NewDashboard"]
    apply_formulas(ws)
    wb.save(path)


if __name__ == "__main__":
    main()
