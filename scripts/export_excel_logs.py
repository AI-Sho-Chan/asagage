import argparse
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook


def dump_sheet(ws, max_rows: int = 10000) -> pd.DataFrame:
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    header = [str(c) if c is not None else f"col{i+1}" for i, c in enumerate(rows[0])]
    data = rows[1:max_rows]
    return pd.DataFrame(data, columns=header)


def export_logs(xlsm_path: Path, outdir: Path) -> List[Path]:
    outdir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(str(xlsm_path), read_only=True, keep_vba=True, data_only=True)
    written: List[Path] = []
    for name in ["Orders", "PnL", "ExecMon", "NewDashboard"]:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        df = dump_sheet(ws)
        if df.empty:
            continue
        path = outdir / f"{name}.csv"
        df.to_csv(path, index=False, encoding="utf-8-sig")
        written.append(path)
    return written


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default=r"C:/AI/asagake/SHINSOKU.xlsm", help="Path to SHINSOKU workbook")
    ap.add_argument("--outdir", default=r"output/trade_logs/latest", help="Destination directory for CSV logs")
    args = ap.parse_args()
    written = export_logs(Path(args.excel), Path(args.outdir))
    print("written:")
    for p in written:
        print(str(p))


if __name__ == "__main__":
    main()

