import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


LEVELS = 10  # top 10 depth on each side


def name(ws: Worksheet, cell: str, name: str) -> None:
    ws.parent.create_named_range(name, ws, cell)


def build_board_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Board"

    # Header
    ws["A1"] = "BoardLogger (Rakuten RSS)"
    ws["A2"] = "Ticker"; ws["B2"] = ""  # value cell
    name(ws, "B2", "Ticker")

    # Columns: BidPx, BidQty, AskPx, AskQty for levels 0..9
    ws["A4"] = "Level"
    ws["B4"] = "BidPx"
    ws["C4"] = "BidQty"
    ws["D4"] = "AskPx"
    ws["E4"] = "AskQty"

    for i in range(LEVELS):
        r = 5 + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=None)
        ws.cell(row=r, column=3, value=None)
        ws.cell(row=r, column=4, value=None)
        ws.cell(row=r, column=5, value=None)
        name(ws, f"B{r}", f"BID_P_{i}")
        name(ws, f"C{r}", f"BID_Q_{i}")
        name(ws, f"D{r}", f"ASK_P_{i}")
        name(ws, f"E{r}", f"ASK_Q_{i}")

    # Aggregates
    ws["G4"] = "Top3Amt"
    ws["H4"] = "Top10Amt"
    ws["G5"] = None
    ws["H5"] = None
    name(ws, "G5", "TOP3_AMT")
    name(ws, "H5", "TOP10_AMT")

    # Placeholders: formulas to be inserted later by install script
    ws["A8"] = "Note: Populate RSS formulas via install_rss_formulas.py when field keys are confirmed."


def build_meta_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Meta")
    ws["A1"] = "This workbook is a safe template for board snapshots."
    ws["A2"] = "Named cells: Ticker, BID_P_0..9, BID_Q_0..9, ASK_P_0..9, ASK_Q_0..9, TOP3_AMT, TOP10_AMT"
    ws["A3"] = "Rakuten RSS functions will be written by install script to these named cells."


def create_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_board_sheet(wb)
    build_meta_sheet(wb)
    wb.save(str(path))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=r"excel/BoardLogger.xlsx")
    args = ap.parse_args()
    create_template(Path(args.out))
    print("written:", args.out)


if __name__ == "__main__":
    main()

